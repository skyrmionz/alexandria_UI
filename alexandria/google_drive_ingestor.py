import os
import io
import json
import argparse
import traceback
import uuid

from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from qdrant_client import QdrantClient
from qdrant_client.http.models import Distance, VectorParams

from langchain_community.embeddings import HuggingFaceEmbeddings

# Additional imports for file processing
from PyPDF2 import PdfReader

# For web credential usage
from google.oauth2.credentials import Credentials

# Scopes: read-only access to Google Drive files.
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

def authenticate_drive(creds_file='credentials.json'):
    """
    Authenticates with Google Drive using OAuth2.
    Opens a local server flow to allow for user login.
    """
    try:
        with open(creds_file, 'r') as f:
            client_config = json.load(f)
        flow = InstalledAppFlow.from_client_config(client_config, SCOPES)
        creds = flow.run_local_server(port=8080)
        service = build('drive', 'v3', credentials=creds)
        return service
    except Exception as e:
        print("Error during interactive authentication:")
        traceback.print_exc()
        raise

def list_drive_files(service, folder_id=None):
    """
    Lists files in the specified folder (non-recursive).
    """
    try:
        query = f"'{folder_id}' in parents" if folder_id else None
        results = service.files().list(
            q=query,
            fields="nextPageToken, files(id, name, mimeType)"
        ).execute()
        items = results.get('files', [])
        print(f"Found {len(items)} files in folder {folder_id}.")
        return items
    except Exception as e:
        print("Error listing files:")
        traceback.print_exc()
        raise

def list_files_recursive(service, folder_id):
    """
    Recursively lists files in a folder and its subfolders.
    Returns a flat list of non-folder files.
    """
    files = list_drive_files(service, folder_id)
    all_files = []
    for f in files:
        if f["mimeType"] == "application/vnd.google-apps.folder":
            print(f"Found folder: {f['name']} (ID: {f['id']}). Crawling its contents...")
            child_files = list_files_recursive(service, f["id"])
            all_files.extend(child_files)
        else:
            all_files.append(f)
    return all_files

def download_file(service, file_id, mime_type):
    """
    Downloads a file. For native Google files (Docs, Sheets, etc.), use export_media
    with an appropriate mime type. For images and other files, use get_media.
    """
    try:
        if mime_type == "application/vnd.google-apps.spreadsheet":
            request = service.files().export_media(fileId=file_id, mimeType="text/csv")
        elif mime_type.startswith("application/vnd.google-apps"):
            request = service.files().export_media(fileId=file_id, mimeType="text/plain")
        else:
            request = service.files().get_media(fileId=file_id)
    
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            print(f"Downloading... {int(status.progress() * 100)}%")
        fh.seek(0)
        return fh
    except Exception as e:
        print(f"Error downloading file {file_id}:")
        traceback.print_exc()
        raise

def extract_text(file_stream, mime_type):
    """
    Extracts text from the file stream.
    - For PDFs: uses PyPDF2.
    - For images: uses OCR via pytesseract.
    - For Excel files: uses openpyxl to extract text.
    - For other files: decodes as text.
    """
    text = ""
    try:
        if mime_type == 'application/pdf':
            reader = PdfReader(file_stream)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        elif mime_type.startswith("image/"):
            from PIL import Image
            import pytesseract
            try:
                image = Image.open(file_stream)
                text = pytesseract.image_to_string(image)
            except pytesseract.pytesseract.TesseractNotFoundError:
                print("Error: Tesseract is not installed or not in PATH. Skipping image file.")
                text = ""
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            # Handle Excel XLSX files using openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(file_stream, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text += str(cell) + " "
                    text += "\n"
        else:
            # Attempt to decode as text using UTF-8 (with replacement for errors)
            text = file_stream.read().decode('utf-8', errors='replace')
    except Exception as e:
        print(f"Error extracting text for mime type {mime_type}:")
        traceback.print_exc()
    return text

def ingest_drive_files(folder_id=None, creds_data=None):
    """
    Authenticates with Google Drive (or uses provided credentials),
    downloads files (from a specified folder and its subfolders),
    vectorizes their content using HuggingFaceEmbeddings,
    and uploads them into Qdrant.
    Supports PDFs, plain text, Google Docs, Google Sheets, Excel files, and images.
    """
    if not folder_id:
        print("No folder ID provided. Please provide a valid folder ID to ingest files from a specific folder.")
        return

    try:
        if creds_data:
            creds = Credentials.from_authorized_user_info(info=creds_data, scopes=SCOPES)
            if not creds.refresh_token:
                print("Error: Credentials do not include a refresh token. Please reauthenticate.")
                return
            drive_service = build('drive', 'v3', credentials=creds)
        else:
            drive_service = authenticate_drive()
    except Exception as e:
        print("Error building Google Drive service:")
        traceback.print_exc()
        return

    try:
        files = list_files_recursive(drive_service, folder_id)
    except Exception as e:
        print("Error during file listing.")
        return

    if not files:
        print("No files found.")
        return

    try:
        qdrant_client = QdrantClient(
            host=os.environ.get("QDRANT_HOST", "localhost"),
            port=int(os.environ.get("QDRANT_PORT", 6333))
        )
        collection_name = "google_drive_docs"
        existing = qdrant_client.get_collections().collections
        if collection_name not in [col.name for col in existing]:
            qdrant_client.recreate_collection(
                collection_name=collection_name,
                vectors_config=VectorParams(size=1024, distance=Distance.COSINE)
            )
            print(f"Created collection '{collection_name}' in Qdrant.")
    except Exception as e:
        print("Error setting up Qdrant:")
        traceback.print_exc()
        return

    try:
        embeddings = HuggingFaceEmbeddings(model_name="intfloat/multilingual-e5-large")
    except Exception as e:
        print("Error initializing embeddings model:")
        traceback.print_exc()
        return

    for file in files:
        print(f"Processing file: {file['name']} (ID: {file['id']}, Type: {file['mimeType']})")
        try:
            file_stream = download_file(drive_service, file['id'], file['mimeType'])
        except Exception as e:
            print(f"Skipping file {file['name']} due to download error.")
            continue

        try:
            text = extract_text(file_stream, file['mimeType'])
        except Exception as e:
            print(f"Skipping file {file['name']} due to text extraction error.")
            continue

        if not text.strip():
            print(f"No text extracted from file {file['name']}; skipping.")
            continue

        try:
            vector = embeddings.embed_query(text)
        except Exception as e:
            print(f"Error vectorizing file {file['name']}:")
            traceback.print_exc()
            continue

        try:
            point_id = str(uuid.uuid5(uuid.NAMESPACE_URL, file['id']))
            qdrant_client.upsert(
                collection_name=collection_name,
                points=[{
                    "id": point_id,
                    "vector": vector,
                    "payload": {
                        "source": file['name'],
                        "text": text
                    }
                }]
            )
            print(f"Uploaded file '{file['name']}' to Qdrant.")
        except Exception as e:
            print(f"Error uploading file '{file['name']}' to Qdrant:")
            traceback.print_exc()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest Google Drive files into Qdrant")
    parser.add_argument('--folder_id', type=str, help='Google Drive Folder ID to ingest (optional)', default=None)
    args = parser.parse_args()
    ingest_drive_files(args.folder_id)